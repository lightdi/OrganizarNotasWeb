import tkinter as tk
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from tkinter import filedialog, messagebox

def extrair_notas(nota):
    a1 = a2 = a3 = a4 = a5 = rec = media = None
    if pd.notna(nota):
        #pattern = r'A1:(\d+|[-]);.*?A2:(\d+|[-]);.*?(?:A3:(\d+|[-]);)?(?:A4:(\d+|[-]);)?(?:A5:(\d+|[-]);)?RE:(\d+|[-]);.*?Media:(\d+)'
        pattern = r'[A-Z]\d:(\d+|[-]);.*?[A-Z]\d:(\d+|[-]);.*?(?:[A-Z]\d:(\d+|[-]);)?(?:[A-Z]\d:(\d+|[-]);)?(?:[A-Z]\d:(\d+|[-]);)?RE:(\d+|[-]);.*?Media:(\d+)'
        matches = re.findall(pattern, nota)

        if matches:
            # Se correspondências são encontradas, atualiza as variáveis
            a1, a2, a3, a4, a5, rec, media = matches[0]

            # Converte '-' para 0
            a1 = '0' if a1 == '-' or a1 == ''else a1
            a2 = '0' if a2 == '-' or a2 == ''else a2
            a3 = '0' if a3 == '-' or a3 == '' else a3
            a4 = '0' if a4 == '-' or a4 == '' else a4
            a5 = '0' if a5 == '-' or a5 == '' else a5
            rec = '0' if rec == '-' or rec == '' else rec
            media = '0' if media == '-' or media == '' else media
            try:
                a = int(a1)
            except e as Exception:
                print (e)


    return int(a1), int(a2), int(a3), int(a4), int(a5), int(rec), int(media)

# Função para selecionar um arquivo e processar
def selecionar_arquivo():
    # Abrir o diálogo para selecionar um único arquivo
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if not file_path:
        messagebox.showinfo("Nenhum arquivo selecionado", "Por favor, selecione um arquivo.")
        return
    
    dfs = pd.read_excel(file_path, sheet_name=None)

    df = dfs['Planilha1']

    dados_processados = []

    for idx, linha in df.iterrows():
        indice = idx
        matricula = str(linha["Matricula"])
        nome = linha["Nome"]
        disciplina = linha['Disciplina']
        situacao = linha['Situacao']
        calculo_nota = linha['Calculo da Nota']
        nota = linha['Nota']
        a1, a2, a3, a4, a5, rec, media = extrair_notas(nota)
        
        conceito = ""


        # Adiciona a linha processada à lista
        dados_processados.append([
            indice,matricula, nome, disciplina, situacao, calculo_nota, a1, a2, a3, a4, a5, rec, media,conceito
        ])

    # Cria um novo DataFrame com os dados processados
    result = pd.DataFrame(dados_processados, columns=[
        '#', 'Matrícula', 'Nome', 'Disciplina', 'Situação', 'Calculo da Nota', 'A1', 'A2', 'A3', 'A4', 'A5', 'REC', 'Média','Conceito'
    ])
    
    output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    

    if output_path:
        try:
            result.to_excel(output_path,index=False)
             # Carregar o arquivo Excel existente
            workbook = load_workbook(output_path)
            sheet = workbook.active  # Selecionar a primeira planilha

            # Criar um preenchimento vermelho para a formatação condicional
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            formula_red = 'M2<70'
            red_rule = FormulaRule( formula=[formula_red], fill=red_fill)
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")    
            formula_green = 'M2>=70'
            grenn_rule = FormulaRule(formula=[formula_green], fill=green_fill)

            # Adiciona as regras de formatação condicional às células da coluna M (supondo que as notas estejam na coluna M)
            sheet.conditional_formatting.add('N2:N{}'.format(sheet.max_row), red_rule)
            sheet.conditional_formatting.add('N2:N{}'.format(sheet.max_row), grenn_rule)

             # Adicionar fórmula SE na coluna 'L'
            for row in range(2, sheet.max_row + 1):
                sheet[f'N{row}'] = f'=IF(M{row}>=70,"Aprovado","Reprovado")'

            workbook.save(output_path)


            messagebox.showinfo("Sucesso",f"Arquivo foram mescaldos com sucesso em {output_path}")
        except Exception as e:
            messagebox.showerror("Erro ao Salvar", f"Erro ao Salvar o arquivo: \n {e}")
    

root= tk.Tk()

root.title("Mesclar Arquivos Excel")


btn_selecionar = tk.Button(root, text="Selecionar arquivos", command=selecionar_arquivo)

btn_selecionar.pack(pady=20)

root.geometry("200x100")
root.resizable(False, False)

root.mainloop()